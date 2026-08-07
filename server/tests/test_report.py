"""Area schedules and the Excel workbook."""
from __future__ import annotations

import io

from openpyxl import load_workbook

from app.plan_schema import (Electrical, Floor, Furniture, Layer, Opening,
                             Project, Room, Vec, Wall)
from app.report import PING_PER_M2, build_report, build_workbook


def room(name: str, w: float, h: float, poly=None) -> Room:
    return Room(id="r_" + name, kind="room", layer="rooms",
                x=0, y=0, w=w, h=h, name=name,
                poly=[Vec(**p) for p in poly] if poly else None)


def furn(item: str, label: str = "") -> Furniture:
    return Furniture(id="f", kind="furniture", layer="furniture",
                     item=item, label=label or item,
                     x=0, y=0, w=10, h=10, angle=0)


def plan(*floors: Floor) -> Project:
    return Project(schemaVersion=1, id="p", name="測試案", layers=[Layer(id="rooms", name="房間",
                   visible=True, locked=False, color="#fff")],
                   floors=list(floors), activeFloorId=floors[0].id if floors else "")


def floor(name: str, objects: list) -> Floor:
    return Floor(id="f_" + name, name=name, elevation=0, height=280, objects=objects)


def test_room_area_uses_the_polygon_when_present():
    poly = [{"x": 0, "y": 0}, {"x": 400, "y": 0}, {"x": 400, "y": 300}, {"x": 0, "y": 300}]
    r = build_report(plan(floor("1F", [room("客廳", 999, 999, poly)])))
    assert abs(r["floors"][0]["rooms"][0]["m2"] - 12) < 1e-9      # polygon wins over w*h


def test_room_area_falls_back_to_the_bounding_box():
    r = build_report(plan(floor("1F", [room("臥室", 400, 300)])))
    assert abs(r["floors"][0]["rooms"][0]["m2"] - 12) < 1e-9


def test_ping_conversion():
    r = build_report(plan(floor("1F", [room("客廳", 400, 300)])))
    got = r["floors"][0]["rooms"][0]
    assert abs(got["ping"] - 12 * PING_PER_M2) < 1e-9
    assert abs(got["ping"] - 3.6299) < 1e-3        # 12 m² ≈ 3.63 坪


def test_totals_add_up_across_floors():
    r = build_report(plan(
        floor("1F", [room("客廳", 400, 300), room("廚房", 200, 200)]),
        floor("2F", [room("主臥", 500, 400)]),
    ))
    assert abs(r["floors"][0]["roomTotalM2"] - (12 + 4)) < 1e-9
    assert abs(r["floors"][1]["roomTotalM2"] - 20) < 1e-9
    assert abs(r["totalM2"] - 36) < 1e-9
    assert abs(r["totalPing"] - 36 * PING_PER_M2) < 1e-9


def test_zero_area_rooms_are_dropped():
    r = build_report(plan(floor("1F", [room("空", 0, 0), room("客廳", 400, 300)])))
    assert [x["name"] for x in r["floors"][0]["rooms"]] == ["客廳"]


def test_furniture_is_counted_by_label():
    r = build_report(plan(floor("1F", [furn("chair", "餐椅"), furn("chair", "餐椅"), furn("sofa", "沙發")])))
    assert r["floors"][0]["furniture"] == [
        {"item": "沙發", "count": 1}, {"item": "餐椅", "count": 2},
    ]


def test_object_counts_cover_every_kind():
    objects = [room("客廳", 400, 300), furn("sofa"),
               Wall(id="w", kind="wall", layer="walls",
                    a=Vec(x=0, y=0), b=Vec(x=100, y=0), thickness=12),
               Opening(id="d", kind="door", layer="openings",
                       x=50, y=0, width=90, angle=0)]
    counts = build_report(plan(floor("1F", objects)))["floors"][0]["counts"]
    assert counts["wall"] == 1 and counts["door"] == 1
    assert counts["furniture"] == 1 and counts["room"] == 1
    assert counts["window"] == 0


def test_an_empty_plan_reports_zeroes():
    r = build_report(plan(floor("1F", [])))
    assert r["totalM2"] == 0
    assert r["floors"][0]["rooms"] == []


def test_workbook_has_a_summary_and_one_sheet_per_floor():
    data = build_workbook(plan(
        floor("1F", [room("客廳", 400, 300)]),
        floor("2F", [room("主臥", 500, 400)]),
    ))
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["總表", "1F", "2F"]
    assert wb["總表"]["B1"].value == "測試案"
    assert wb["1F"]["A1"].value == "房間"
    assert wb["1F"]["A2"].value == "客廳"
    assert abs(wb["1F"]["B2"].value - 12) < 1e-9


def test_workbook_sheet_names_are_made_legal_for_excel():
    data = build_workbook(plan(floor("1F/B1*地下[室]", [room("客廳", 400, 300)])))
    wb = load_workbook(io.BytesIO(data))
    name = wb.sheetnames[1]
    assert not set(name) & set("[]:*?/\\")
    assert len(name) <= 31


def test_workbook_survives_a_plan_with_no_rooms():
    wb = load_workbook(io.BytesIO(build_workbook(plan(floor("1F", [furn("sofa")])))))
    assert "1F" in wb.sheetnames


def elec(item: str, label: str = "") -> Electrical:
    return Electrical(id="e", kind="electrical", layer="electrical",
                      item=item, label=label or item, x=0, y=0, angle=0)


def test_electrical_fittings_are_counted_by_type():
    r = build_report(plan(floor("1F", [
        elec("socket2", "雙插座"), elec("socket2", "雙插座"),
        elec("switch1", "單切開關"), elec("downlight", "崁燈"),
    ])))
    assert r["floors"][0]["electrical"] == [
        {"item": "單切開關", "count": 1},
        {"item": "崁燈", "count": 1},
        {"item": "雙插座", "count": 2},
    ]
    assert r["floors"][0]["counts"]["electrical"] == 4


def test_the_workbook_lists_the_electrical_schedule():
    data = build_workbook(plan(floor("1F", [room("客廳", 400, 300), elec("socket2", "雙插座")])))
    wb = load_workbook(io.BytesIO(data))
    values = [c.value for row in wb["1F"].iter_rows(values_only=True) for c in
              [type("C", (), {"value": v})() for v in row]]
    assert "水電配置" in values
    assert "雙插座" in values


def test_a_plan_with_no_fittings_omits_the_electrical_block():
    r = build_report(plan(floor("1F", [room("客廳", 400, 300)])))
    assert r["floors"][0]["electrical"] == []
