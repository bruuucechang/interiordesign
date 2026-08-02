"""Area schedules and the Excel workbook built from them.

The editor already knows every room's polygon and every piece of furniture, but
had no way to get a summary out of it. This turns a saved plan into the tally a
quote is written from.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .rooms import polygon_signed_area

PING_PER_M2 = 1 / 3.30579   # 1 坪 = 3.30579 m²


def _area_m2(room: dict[str, Any]) -> float:
    poly = room.get("poly")
    if poly and len(poly) >= 3:
        return abs(polygon_signed_area(poly)) / 10000
    return float(room.get("w", 0)) * float(room.get("h", 0)) / 10000


def build_report(project: dict[str, Any]) -> dict[str, Any]:
    """Per-floor room and furniture tallies, plus a project total."""
    floors_out = []
    for floor in project.get("floors") or []:
        objects = floor.get("objects") or []

        rooms = []
        for o in objects:
            if o.get("kind") != "room":
                continue
            m2 = _area_m2(o)
            if m2 <= 0:
                continue
            rooms.append({"name": o.get("name") or "房間", "m2": m2, "ping": m2 * PING_PER_M2})

        counts: dict[str, int] = {}
        for o in objects:
            if o.get("kind") == "furniture":
                label = o.get("label") or o.get("item") or "家具"
                counts[label] = counts.get(label, 0) + 1
        furniture = [{"item": k, "count": v} for k, v in sorted(counts.items())]

        tally = {
            kind: sum(1 for o in objects if o.get("kind") == kind)
            for kind in ("wall", "beam", "door", "window", "furniture", "room")
        }

        floors_out.append(
            {
                "id": floor.get("id"),
                "name": floor.get("name") or "",
                "rooms": rooms,
                "roomTotalM2": sum(r["m2"] for r in rooms),
                "roomTotalPing": sum(r["ping"] for r in rooms),
                "furniture": furniture,
                "counts": tally,
            }
        )

    return {
        "project": project.get("name") or "未命名平面圖",
        "floors": floors_out,
        "totalM2": sum(f["roomTotalM2"] for f in floors_out),
        "totalPing": sum(f["roomTotalPing"] for f in floors_out),
    }


# ---------------------------------------------------------------- workbook

_HEAD_FILL = PatternFill("solid", fgColor="1F2937")
_HEAD_FONT = Font(color="FFFFFF", bold=True)
_THIN = Side(style="thin", color="B0B7C3")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_row(ws, row: int, values: list[Any], *, head: bool = False, bold: bool = False) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = _BORDER
        if head:
            cell.fill = _HEAD_FILL
            cell.font = _HEAD_FONT
        elif bold:
            cell.font = Font(bold=True)
        if isinstance(value, float):
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="right")


def build_workbook(project: dict[str, Any]) -> bytes:
    """The report as a .xlsx: one sheet per floor, plus a summary."""
    report = build_report(project)
    wb = Workbook()

    summary = wb.active
    summary.title = "總表"
    _write_row(summary, 1, ["專案", report["project"]], bold=True)
    _write_row(summary, 3, ["樓層", "房間數", "面積 (m²)", "面積 (坪)"], head=True)
    row = 4
    for f in report["floors"]:
        _write_row(summary, row, [f["name"], len(f["rooms"]), f["roomTotalM2"], f["roomTotalPing"]])
        row += 1
    _write_row(summary, row, ["合計", sum(len(f["rooms"]) for f in report["floors"]),
                              report["totalM2"], report["totalPing"]], bold=True)

    for f in report["floors"]:
        # Excel forbids []:*?/\ in sheet names and caps them at 31 chars.
        title = (f["name"] or "樓層").translate({ord(c): "-" for c in "[]:*?/\\"})[:31]
        ws = wb.create_sheet(title or "樓層")
        _write_row(ws, 1, ["房間", "面積 (m²)", "面積 (坪)"], head=True)
        r = 2
        for room in f["rooms"]:
            _write_row(ws, r, [room["name"], room["m2"], room["ping"]])
            r += 1
        _write_row(ws, r, ["合計", f["roomTotalM2"], f["roomTotalPing"]], bold=True)

        r += 2
        _write_row(ws, r, ["家具", "數量"], head=True)
        r += 1
        for item in f["furniture"]:
            _write_row(ws, r, [item["item"], item["count"]])
            r += 1

        r += 1
        _write_row(ws, r, ["物件統計", ""], head=True)
        r += 1
        labels = {"wall": "牆", "beam": "樑", "door": "門",
                  "window": "窗", "furniture": "家具", "room": "房間"}
        for kind, label in labels.items():
            _write_row(ws, r, [label, f["counts"].get(kind, 0)])
            r += 1

    for ws in wb.worksheets:
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
